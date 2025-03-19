import openpyxl
from openpyxl import load_workbook

def process_excel_file(file_name="C:\\Users\\Arotiana\\Documents\\Scrap\\Hugo_Boss\\HugoBoss.xlsx"):
    """
    Process an Excel file using column numbering instead of column letters.
    
    Args:
        file_name (str): Name of the Excel file in the current directory
    """
    # Load the workbook
    print(f"Loading workbook: {file_name}")
    wb = load_workbook(file_name)
    ws = wb.active
    
    # Define column numbers (AD=30, Q=17, AB=28, AE=31)
    fit_check_col = 30    # Column AD
    fit_store_col = 17    # Column Q
    fill_target_col = 28  # Column AB
    fill_source_col = 31  # Column AE
    
    # Track counters for reporting
    fit_replacements = 0
    filled_values = 0
    
    # Iterate through each row in the worksheet
    for row_idx in range(1, ws.max_row + 1):
        # Step 1 & 2: Check if cell in AD contains fit values
        cell_AD = ws.cell(row=row_idx, column=fit_check_col)
        cell_value = cell_AD.value
        
        if cell_value in ["Slim fit", "Regular fit", "Relaxed fit"]:
            # Store the original fit value in column Q
            cell_Q = ws.cell(row=row_idx, column=fit_store_col)
            cell_Q.value = cell_value
            
            # Replace with "Normal sale" in column AD
            cell_AD.value = "Normal sale"
            fit_replacements += 1
        
        # Step 3: Fill None values in column AB with values from AE where AE doesn't contain "Sale-"
        cell_AB = ws.cell(row=row_idx, column=fill_target_col)
        cell_AE = ws.cell(row=row_idx, column=fill_source_col)
        
        ab_value = cell_AB.value
        ae_value = cell_AE.value
        
        # Check if AB is empty/None and AE doesn't contain "Sale-"
        if (ab_value is None or ab_value == "None" or ab_value == ""):
            if ae_value and isinstance(ae_value, str) and "Sale-" not in ae_value:
                cell_AB.value = ae_value
                filled_values += 1
    
    # Save the modified workbook to a new file
    output_file = "HugoBoss_processed.xlsx"
    wb.save(output_file)
    print(f"Processed {fit_replacements} rows with fit values")
    print(f"Filled {filled_values} empty AB values with AE values")
    print(f"Saved processed file to: {output_file}")
    
    return output_file

if __name__ == "__main__":
    # Process the file
    processed_file = process_excel_file()
    print(f"Processing complete. File saved at: {processed_file}")